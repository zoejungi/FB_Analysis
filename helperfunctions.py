# helperfunctions for main_subj.py, GDI.py or class functions (in Subject Class)

import math
import pandas as pd
import os
import numpy as np
import matplotlib.colors as mcolors
from scipy.stats import pearsonr
from scipy.stats import spearmanr
from openpyxl import load_workbook

def get_filepath (datapath, sub_id, FB, param):
    # input = datapath to all data folders, sub_ID (='S1_', don't forget the underscore), FB = targeted param, param = analyzed param
    # output = filepath to needed file for analyzing different params (APF = DFlowData, ST & other spatiotemporal params = spatiotemp files,
    # POF & everything else = extraction_normalization file)

    if param == 'APF':
        # take DFlow data
        dflow_path = os.path.join (datapath, 'D-FlowData')
        for sub_folder in os.listdir(dflow_path):
            if sub_id in sub_folder:
                sub_path = os.path.join(dflow_path, sub_folder)
                break
        if os.path.exists(sub_path) and os.path.isdir(sub_path):
            for file in os.listdir(sub_path):
                if FB.lower() in file.lower() and 'HBM2' in file:
                    filepath = os.path.join(sub_path, file)
                    break
        else:
            print(f'Error: {param} filepath for {sub_id} during FB targeting {FB} not found')

    elif param == 'spatiotemp' or param == 'ST': # includes ST, step length, swing time, step height, step width etc
        # take spatiotemp data in analysis folder
        ana_path = os.path.join(datapath, 'analysis')
        for sub_folder in os.listdir(ana_path):
            if sub_id in sub_folder:
                sub_path = os.path.join(ana_path, sub_folder)
                break
        if os.path.exists(sub_path) and os.path.isdir(sub_path):
            for file in os.listdir(sub_path):
                if FB.lower() in file.lower() and 'spatiotemp' in file.lower():
                    filepath = os.path.join(sub_path, file)
                    break
        else:
            print(f'Error: {param} filepath for {sub_id} during FB targeting {FB} not found')

    else:
        # go to extraction_normalization folder
        extr_path = os.path.join(datapath, 'extraction_normalization')
        for sub_folder in os.listdir(extr_path):
            if sub_id in sub_folder:
                sub_path = os.path.join(extr_path, sub_folder)
                break
        if os.path.exists(sub_path) and os.path.isdir(sub_path):
            for out_folder in os.listdir(sub_path):
                if FB.lower() in out_folder.lower():
                    out_path = os.path.join(sub_path, out_folder)
                    break
        if os.path.exists(out_path) and os.path.isdir(out_path):
            for file in os.listdir(out_path):
                if FB.lower() in file.lower() and param.lower() in file.lower():
                    filepath = os.path.join(out_path, file)
                    break
        else:
            print(f'Error: {param} filepath for {sub_id} during FB targeting {FB} not found')

    return filepath
def convert_txt_to_csv (input_path, separator = ','):
    # input = input_path = file including its path, separator of txt file
    # output = csv.file

    base, _ = os.path.splitext(input_path)
    output_path = base + '.csv'
    # Read the input text file
    file = pd.read_csv(input_path, sep=separator)
    # Write the DataFrame to a CSV file
    file.to_csv(output_path, index=False)

    return output_path
def identify_maxAPF (df_apf, sub_id=[], datapath=[]):
    # input = df_apf = raw DFlowdata, sub_id and datapath for saving data
    # output = list of dfs (during ST, POF and APF FB), headers include  'apf_left', 'apf_right',
    #             'time', 'real_time_left', 'real_time_right', 'gait_cycle_left', 'gait_cycle_right',
    #             'group_id' (for saving all three datafiles in one file -> 0 = ST, 1 = POF, 2 = APF)

    # Identifying max APF
    df_final = []
    for i, df in enumerate(df_apf):

        group_id = i
        apf_left_i = []
        apf_right_i = []
        real_time_max_apf_left_i = []
        real_time_max_apf_right_i = []
        gait_cycle_left_i = []
        gait_cycle_right_i = []

        prev_lto = 2.0
        prev_rto = 2.0

        for index, row in df_apf[i].iterrows():

            if (row["LHS"] >= 2.0) & (row["LTO"] >= 2.0):
                rto = row["RTO"]
                lto = row["LTO"]

                if rto > prev_rto:  # left heel strike event

                    rto_mask = df_apf[i]["RTO"] == rto  # left mid stance phase start
                    lto_mask = df_apf[i]["LTO"] == lto
                    last_lto_index = df_apf[i][lto_mask].index[-1]
                    next_30_rows_mask = (df_apf[i].index > last_lto_index) & (df_apf[i].index <= last_lto_index + 30)
                    mask = rto_mask & next_30_rows_mask

                    left_max_apf = df_apf[i].loc[mask, "LAPF.Offset"].min()
                    if (not math.isnan(left_max_apf)):  # sometime mask is empty and min = nan -> idxnan dont work
                        min_index_left = df_apf[i].loc[mask, "LAPF.Offset"].idxmin()
                        time_left_max_apf = df_apf[i].loc[min_index_left, "Time_adj"]

                        apf_left_i.append(left_max_apf)
                        # time_apf_i.append()
                        real_time_max_apf_left_i.append(float(time_left_max_apf))
                        # gait_cycle_left_i.append(max(df_apf[i].loc[min_index_left, "LHS"], df_apf[i].loc[min_index_left, "RHS"])) # if you want gait cycle depending of wich foot the subject move first
                        gait_cycle_left_i.append(df_apf[i].loc[min_index_left, "LHS"])  # gait cycle for left foot # I use this one because this is what we did for the other parameters in the other notebook

                if lto > prev_lto:

                    lto_mask = df_apf[i]["LTO"] == lto  # right mid stance phase start
                    rto_mask = df_apf[i]["RTO"] == rto
                    last_rto_index = df_apf[i][rto_mask].index[-1]
                    next_30_rows_mask_right = (df_apf[i].index > last_rto_index) & (df_apf[i].index <= last_rto_index + 30)
                    mask_right = lto_mask & next_30_rows_mask_right

                    right_max_apf = df_apf[i].loc[mask_right, "RAPF.Offset"].min()
                    if (not math.isnan(right_max_apf)):
                        min_index_right = df_apf[i].loc[mask_right, "RAPF.Offset"].idxmin()
                        time_right_max_apf = df_apf[i].loc[min_index_right, "Time_adj"]

                        apf_right_i.append(right_max_apf)
                        real_time_max_apf_right_i.append(float(time_right_max_apf))
                        gait_cycle_right_i.append(df_apf[i].loc[min_index_left, "LHS"])  # gait cycle for left foot # I use this one because this is what we did for the other parameters in the other notebook

                prev_rto = rto
                prev_lto = lto

        unique_event_left = [element for element in gait_cycle_left_i if element not in gait_cycle_right_i]
        unique_event_right = [element for element in gait_cycle_right_i if element not in gait_cycle_left_i]

        if unique_event_right or unique_event_left:
            for index in range(len(unique_event_right)):
                ind = gait_cycle_right_i.index(unique_event_right[index])
                del apf_right_i[ind]
                del real_time_max_apf_right_i[ind]
                del gait_cycle_right_i[ind]

            for index in range(len(unique_event_left)):
                ind = gait_cycle_left_i.index(unique_event_left[index])
                del apf_left_i[ind]
                del real_time_max_apf_left_i[ind]
                del gait_cycle_left_i[ind]

        df_final_i = pd.DataFrame({
            'apf_left': apf_left_i,
            'apf_right': apf_right_i,
            'time': real_time_max_apf_left_i,
            'real_time_left': real_time_max_apf_left_i,
            'real_time_right': real_time_max_apf_right_i,
            'gait_cycle_left': gait_cycle_left_i,
            'gait_cycle_right': gait_cycle_right_i,
            'group_id': group_id
        })
        df_final.append(df_final_i)
    if sub_id and datapath:
        pd.DataFrame(pd.concat(df_final, ignore_index=True)).to_csv(os.path.join(datapath, f'maxAPF_identified_{sub_id}.csv'), index=False)
    return df_final

def check_cycles (df):
    # input = df_apf (should have gait_cycle_left and gait_cycle_right)
    # output = True (if no event is missing), False (if event is missing)
    # to check that no event is missing for one foot and not the other
    for i, data in enumerate(df):
        unique_left = [element for element in df[i]['gait_cycle_left'] if element not in df[i]['gait_cycle_right']]
        unique_right = [element for element in df[i]['gait_cycle_right'] if element not in df[i]['gait_cycle_left']]
        print('checked')
        if unique_left or unique_right:
            print(unique_left)
            print(unique_right)
            print("Error: there's at least one event missing in either foot")
            return False
    return True
def exclude_outliers(df):
    # input = list of dataframes -> column headers needed: 'apf_left', 'apf_right'
    # output = same list of dataframes without all the outliers (apf>=-50, apf>=20 or apf=None)

    # Iterate over each DataFrame in the list
    for i in range(len(df)):
        # Create a boolean mask to identify the rows to keep
        mask = ((-50 <= df[i]['apf_left']) & (df[i]['apf_left'] <= 20) & (df[i]['apf_left'] != None) &
                (-50 <= df[i]['apf_right']) & (df[i]['apf_right'] <= 20) & (df[i]['apf_right'] != 0) & (df[i]['apf_right'] != None))

        print("Original length:", len(df[i]['apf_left']))
        # Apply the mask to filter the DataFrame and reset the index
        df[i] = df[i][mask].reset_index(drop=True)
        print("New length:", len(df[i]['apf_left']))
def print_in_excel_table(value, sheet_name, row_header, column_header, output_file):
    # input = value to print, sheet_name, row_header (needs to be already created), column_header (needs to be already created), path to output-file
    # output = None -> only printing (excel needs to be closed for printing, otherwise it's not working)

    # Load the existing Excel file
    book = load_workbook(output_file)

    # Check if the sheet exists
    if sheet_name not in book.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        return

    # Load the sheet into a DataFrame
    df = pd.read_excel(output_file, sheet_name=sheet_name, index_col=0)

    # Find the row and column index based on the row and column headers
    try:
        row_index = df.index.get_loc(row_header)
    except KeyError:
        print(f"Row header '{row_header}' not found.")
        return
    try:
        column_index = df.columns.get_loc(column_header)
    except KeyError:
        print(f"Column header '{column_header}' not found.")
        return

    # Set value in specific cell
    df.iloc[row_index, column_index] = value

    # Write the DataFrame back to the same sheet in the Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # No need to set writer.book or writer.sheets explicitly
        df.to_excel(writer, sheet_name=sheet_name)
def calculate_averages (subjects, FB_mode, param):
    # input = subjects = list of subjects to calculate the average on, FB_mode = 'hFB' or 'vFB', param = analyzed param all inlower case
    # output = 3 pd.series (mean_SR, std_SR, mean_time)

    t_intervals = [(0, 60), (60, 240), (240, 300), (300, 480), (480, 660)]
    participant_averages = {}
    if FB_mode.lower() == 'st':
        fb = 0
    elif FB_mode.lower() == 'pof':
        fb = 1
    elif FB_mode.lower() == 'apf':
        fb = 2
    for i, subject in enumerate(subjects.values()):
        t_average = []
        symmetry_ratio = []
        if not (subject.study=='vFB' and (subject.ID =='S2_' or subject.ID=='S3_' or subject.ID=='S4_' or subject.ID== 'S6_' or subject.ID=='S12_')):
            for start_time, end_time in t_intervals:
                filt = (subject.__getattribute__(param)[fb]['time'] >= start_time) & (subject.__getattribute__(param)[fb]['time'] <= end_time)
                df_phase = subject.__getattribute__(param)[fb][filt]
                df_phase.reset_index(drop=True, inplace=True)
                total_datapoints = df_phase.shape[0]
                group_size = int(total_datapoints * 0.1)  # 10% of phase

                # Calculate the average for each group -> 10 groups
                for j in range(10):
                    start_index = j * group_size
                    end_index = (j + 1) * group_size

                    sym_ratios = df_phase.iloc[start_index:end_index + 1, :]['SR'] # SR already divided by baseline SR
                    symmetry_ratio.append(sym_ratios.mean())
                    t_average.append(df_phase.loc[start_index:end_index,['time']].mean())

                    participant_averages[i] = {
                        'time_average': t_average,
                        'symmetry_ratio': symmetry_ratio
                    }

    # Averaging over all participants
    mean_SR = []
    std_SR = []
    mean_time = []
    # Assuming each participant has the same number of entries in 'symmetry_ratio' and 'time_average'
    num_entries = len(next(iter(participant_averages.values()))['symmetry_ratio'])
    # Calculate mean and std for each entry position across participants
    for k in range(num_entries):
        symmetry_ratios_at_position = [avg_data['symmetry_ratio'][k] for avg_data in participant_averages.values()]
        time_at_position = [avg_time['time_average'][k] for avg_time in participant_averages.values()]
        mean_SR.append(np.mean(symmetry_ratios_at_position))
        std_SR.append(np.mean(np.std(symmetry_ratios_at_position) / np.sqrt(len(symmetry_ratios_at_position))))
        mean_time.append(np.mean(time_at_position))

    return pd.Series(mean_SR), pd.Series(std_SR), pd.Series(mean_time)
def get_color(base_color, incr):
    #used for getting shades of the same color for the graphs

    base_rgb = mcolors.to_rgb(base_color)
    similar_color = tuple(min(1, c + incr) for c in base_rgb)  # Adjust the increment value as needed
    return similar_color

def weighted_std_mean(values, weights):
    # used for calculating means for SUS

    mean = np.average(values, weights=weights)
    variance = np.average((values - mean) ** 2, weights=weights)
    return np.sqrt(variance), mean

def calculate_mean_interval(t_start, t_end, df, decimals = 2):
    #used to calculate mean SR in a defined interval (i.e. individual subjects during each FB phase)

    filt = (df['time'] >= t_start) & (df['time'] <= t_end)
    df_filt = df[filt]
    mean = np.mean(df_filt['SR'])
    return np.round(mean,decimals)

def group_data(subject, FB, ID, param):
    # input = subject for correlation (only 1), FB (mode), ID, param analysed
    # output = dataframe for correlation for specific subject

    if FB == 'ST':
        j = 0
    elif FB == 'POF':
        j = 1
    elif FB == 'APF':
        j = 2
    if param == 'apf':
        df_NW = subject.__getattribute__(param)[j][subject.__getattribute__(param)[j]['time'] <= 60]
        df_NW = df_NW[['gait_cycle_left', 'SR_raw']]
        df_NW['condition'] = f'NWduring{FB}'
        df_NW['subject_ID'] = ID
        df_FB2 = subject.__getattribute__(param)[j][(420 <= subject.__getattribute__(param)[j]['time']) & (subject.__getattribute__(param)[j]['time'] <= 480)]
        df_FB2 = df_FB2[['gait_cycle_left', 'SR_raw']]
        df_FB2['condition'] = f'during{FB}'
        df_FB2['subject_ID'] = ID
    else:
        df_NW = subject.__getattribute__(param)[j][subject.__getattribute__(param)[j]['time'] <= 60]
        df_NW = df_NW[['cycle_number', 'SR_raw']]
        df_NW['condition'] = f'NWduring{FB}'
        df_NW['subject_ID'] = ID
        df_FB2 = subject.__getattribute__(param)[j][(420 <= subject.__getattribute__(param)[j]['time']) & (subject.__getattribute__(param)[j]['time'] <= 480)]
        df_FB2 = df_FB2[['cycle_number', 'SR_raw']]
        df_FB2['condition'] = f'during{FB}'
        df_FB2['subject_ID'] = ID
    return pd.concat([df_NW, df_FB2], ignore_index = True)

def calc_stats (df_corr, FB):
    # input = correlation dataframe with all subjects included, all conditions ('NWduringST', 'duringST', etc), all params)
    # output = meanStats during FB (=dataframe with 1 row and 19 headers)

    mean_stats = {'subject_ID': []}

    columns_to_mean = ["SR_ST", "SR_POF", "SR_APF", "SR_swingtime", "SR_steplength", "SR_stepwidth", "SR_stepheight",
                       "SR_meanGRFz", "SR_kneeflexion"]
    for col in columns_to_mean:
        mean_stats[f'{col}_NW'] = []
        mean_stats[col] = []

    for col in columns_to_mean:
        # Calculate the mean of the last 10 gait cycles for the column where condition is met'
        mean_stats[f'{col}_NW'].append(df_corr[df_corr['condition'] == 'NW'+FB][col].tail(10).mean())
        mean_stats[col].append(df_corr[df_corr['condition'] == FB][col].tail(10).mean())
    mean_stats['subject_ID'] = df_corr['subject_ID'][1]
    return pd.DataFrame(mean_stats)

def save_corrstats(df, path, FB=False):
    # input = df (correlation or stats, if stats then FB = 'ST', 'POF', or 'APF'), path = saving path
    # output = None

    df = pd.concat(df)
    if FB == False:
        df = df.sort_values(by=['condition', 'subject_ID', 'cycle_number'])
        df.reset_index(drop=True, inplace=True)
        df.to_csv(path, index=False)
    else:
        stats_path = os.path.join(path, 'stats' + FB + '.csv')
        df.to_csv(stats_path, index=False)
def calc_corrcoeffs(df, FB_mode, output_file):
    # input = correlation dataframe, FB_mode, output_file to save data
    # output = None, only printing in excel.

    df_ST = df[(df['condition'] == 'NWduringST') | (df['condition'] == 'duringST')]
    df_POF = df[(df['condition'] == 'NWduringPOF') | (df['condition'] == 'duringPOF')]
    df_APF = df[(df['condition'] == 'NWduringAPF') | (df['condition'] == 'duringAPF')]

    columns_ST = df_ST.columns.drop(['condition', 'subject_ID', 'cycle_number', 'SR_ST'])
    for i in range(len(columns_ST)):
        pearson_ST = pearsonr(df_ST['SR_ST'], df_ST[columns_ST[i]])[0]
        spearman_ST = spearmanr(df_ST['SR_ST'], df_ST[columns_ST[i]])[0]
        print_in_excel_table(pearson_ST, 'Correlation', FB_mode + ' pearsons', 'STvs'+columns_ST[i], output_file)
        print_in_excel_table(spearman_ST, 'Correlation', FB_mode + ' spearmans', 'STvs' + columns_ST[i], output_file)

    columns_POF = df_POF.columns.drop(['condition', 'subject_ID', 'cycle_number', 'SR_POF'])
    for i in range(len(columns_POF)):
        pearson_POF = pearsonr(df_POF['SR_POF'], df_POF[columns_POF[i]])[0]
        spearman_POF = spearmanr(df_POF['SR_POF'], df_POF[columns_POF[i]])[0]
        print_in_excel_table(pearson_POF, 'Correlation', FB_mode + ' pearsons', 'POFvs' + columns_POF[i], output_file)
        print_in_excel_table(spearman_POF, 'Correlation', FB_mode + ' spearmans', 'POFvs' + columns_POF[i], output_file)

    columns_APF = df_APF.columns.drop(['condition', 'subject_ID', 'cycle_number', 'SR_APF'])
    for i in range(len(columns_APF)):
        pearson_APF = pearsonr(df_APF['SR_APF'], df_APF[columns_APF[i]])[0]
        spearman_APF = spearmanr(df_APF['SR_APF'], df_APF[columns_APF[i]])[0]
        print_in_excel_table(pearson_APF, 'Correlation', FB_mode + ' pearsons', 'APFvs' + columns_APF[i], output_file)
        print_in_excel_table(spearman_APF, 'Correlation', FB_mode + ' spearmans', 'APFvs' + columns_APF[i], output_file)
